# -*- coding:utf-8 -*-
from flask import Flask
from flask_restful import Api, Resource, abort, reqparse
from app import db
from app.models import ConcGateway, ConcNode, ConcTemp
from sqlalchemy import and_
from flasgger import Swagger, swag_from
import random
from utils import index_color
import datetime
import generate_report
from openpyxl import load_workbook
from openpyxl import Workbook

app = Flask(__name__)
api = Api(app)

swagger = Swagger(app)


app.config['SWAGGER'] = {
    'title': 'SmartLinkCloud RESTful API',
    'uiversion': 2
}
# swag = Swagger(app)


TODOS = {
    'todo1': {'task': 'build an API'},
    'todo2': {'task': '?????'},
    'todo3': {'task': 'profit!'},
    '42': {'task': 'Use Flasgger'}
}

def abort_if_todo_doesnt_exist(todo_id):
    if todo_id not in TODOS:
        abort(404, message="Todo {} doesn't exist".format(todo_id))

parser = reqparse.RequestParser()
parser.add_argument('task')


# Todo
# shows a single todo item and lets you delete a todo item
class Todo(Resource):
    def get(self, todo_id):
        """
        This is an example
        ---
        tags:
          - restful
        parameters:
          - in: path
            name: todo_id
            required: true
            description: The ID of the task, try 42!
            type: string
        responses:
          200:
            description: The task data
            schema:
              id: Task
              properties:
                task:
                  type: string
                  default: My Task
        """
        abort_if_todo_doesnt_exist(todo_id)
        return TODOS[todo_id]

    def delete(self, todo_id):
        """
        This is an example
        ---
        tags:
          - restful
        parameters:
          - in: path
            name: todo_id
            required: true
            description: The ID of the task, try 42!
            type: string
        responses:
          204:
            description: Task deleted
        """
        abort_if_todo_doesnt_exist(todo_id)
        del TODOS[todo_id]
        return '', 204

    def put(self, todo_id):
        """
        This is an example
        ---
        tags:
          - restful
        parameters:
          - in: body
            name: body
            schema:
              $ref: '#/definitions/Task'
          - in: path
            name: todo_id
            required: true
            description: The ID of the task, try 42!
            type: string
        responses:
          201:
            description: The task has been updated
            schema:
              $ref: '#/definitions/Task'
        """
        args = parser.parse_args()
        task = {'task': args['task']}
        TODOS[todo_id] = task
        return task, 201


# TodoList
# shows a list of all todos, and lets you POST to add new tasks
class TodoList(Resource):
    def get(self):
        """
        This is an example
        ---
        tags:
          - restful
        responses:
          200:
            description: The task data
            schema:
              id: Tasks
              properties:
                task_id:
                  type: object
                  schema:
                    $ref: '#/definitions/Task'
        """
        return TODOS

    def post(self):
        """
        This is an example
        ---
        tags:
          - restful
        parameters:
          - in: body
            name: body
            schema:
              $ref: '#/definitions/Task'
        responses:
          201:
            description: The task has been created
            schema:
              $ref: '#/definitions/Task'
        """
        args = parser.parse_args()
        todo_id = int(max(TODOS.keys()).lstrip('todo')) + 1
        todo_id = 'todo%i' % todo_id
        TODOS[todo_id] = {'task': args['task']}
        return TODOS[todo_id], 201


class Username(Resource):
    @swag_from('username_specs.yml', methods=['GET'])
    def get(self, username):
        return {'username': username}, 200


class ConcRealtimeTemp(Resource):

    def get(self):

        parser = reqparse.RequestParser()
        parser.add_argument('gatewayAddr', type=str)
        parser.add_argument('nodeAddr', type=str)

        args = parser.parse_args()

        print('-------ConcRealtimeTemp args---------', args)

        gatewayAddr = args['gatewayAddr']
        nodeAddr = args['nodeAddr']

        temps = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5, ConcTemp.temp6, ConcTemp.battery_vol, 
            ConcTemp.conc_node_id).filter(ConcTemp.conc_node_id == nodeAddr).order_by(ConcTemp.datetime.desc()).first()
        # temps = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5, ConcTemp.temp6, ConcTemp.battery_vol, 
        #     ConcNode.node_addr).join(ConcNode, ConcNode.id == ConcTemp.conc_node_id).filter(ConcNode.node_addr == nodeAddr).order_by(ConcTemp.datetime.desc()).first()

        print("temps:", temps)
        conc_realtime_temp_dic = {"concRealtimeTemp": [{"icon": "bulb", "color": "#64ea91", "title": "温度1", "number": temps[0]}, 
        {"icon": "bulb", "color": "#8fc9fb", "title": "温度2", "number": temps[1]}, {"icon": "bulb", "color": "#d897eb", "title": "温度3", "number": temps[2]}, 
        {"icon": "bulb", "color": "#d897eb", "title": "温度4", "number": temps[3]}, {"icon": "bulb", "color": "#d897eb", "title": "温度5", "number": temps[4]}, 
        {"icon": "bulb", "color": "#d897eb", "title": "温度6", "number": temps[5]}, {"icon": "home", "color": "#f69899", "title": "电池", "number": temps[6]}]}
        return conc_realtime_temp_dic

    def delete(self, todo_id):
        pass

    def put(self, todo_id):
        pass


class ConcTemps(Resource):
    '''
        get the lates 10 temps.
    '''
    def get(self):

        parser = reqparse.RequestParser()
        parser.add_argument('gatewayAddr', type=str)
        parser.add_argument('nodeAddr', type=str)

        args = parser.parse_args()

        print('-------ConcTemps args---------', args)

        gatewayAddr = args['gatewayAddr']
        nodeAddr = args['nodeAddr']

        temp_records = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5, ConcTemp.temp6, ConcTemp.datetime
            ).filter(ConcTemp.conc_node_id == nodeAddr).order_by(ConcTemp.datetime.desc()).limit(10).all()
        # temp_records = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5, ConcTemp.temp6, ConcTemp.datetime
        #     ).join(ConcNode, ConcNode.id == ConcTemp.conc_node_id).filter(ConcNode.node_addr == nodeAddr).order_by(ConcTemp.datetime.desc()).limit(10).all()

        temp_log = []
        for i in xrange(len(temp_records)):
            temp_log.append({"时间": temp_records[i][6].strftime("%Y-%m-%d %H:%M:%S"), "温度1": temp_records[i][0], "温度2": temp_records[i][1], 
                "温度3": temp_records[i][2],"温度4": temp_records[i][3], "温度5": temp_records[i][4], "温度6": temp_records[i][5]})

        temps_reverse = temp_log[::-1]
        print('------------temps_reverse--------------')
        print(temps_reverse)

        temps_dict = {"concTemps": temps_reverse}
        return temps_dict

    def delete(self, todo_id):
        pass

    def put(self, todo_id):
        pass


class ConcTempRecord(Resource):
    '''
        get the temp records by the input datetime. %H:%M:S%
    '''
    def get(self):

        parser = reqparse.RequestParser()
        parser.add_argument('gatewayAddr', type=str)
        parser.add_argument('nodeAddr', type=str)
        parser.add_argument('startTime', type=str)
        parser.add_argument('endTime', type=str)

        args = parser.parse_args()

        print('-------ConcTempRecord args---------', args)

        gatewayAddr = args['gatewayAddr']
        nodeAddr = args['nodeAddr']
        startTime0 = args['startTime']
        endTime0 = args['endTime']

        datetime_list = generate_report.generate_datetime_list(startTime0,endTime0)
        # print(datetime_list)

        startTime = datetime.datetime.strptime(startTime0, "%Y-%m-%d %H:%M:%S")
        endTime = datetime.datetime.strptime(endTime0, "%Y-%m-%d %H:%M:%S")
        print('-------startTime--------')
        print('--------endTime-------')
        print(startTime)
        print(endTime)
        temp_records = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5,
            ConcTemp.temp6, ConcTemp.battery_vol, ConcTemp.datetime,ConcTemp.conc_node_id).filter(and_(ConcTemp.conc_node_id == nodeAddr, ConcTemp.datetime.between(startTime, endTime))).order_by(ConcTemp.datetime.desc()).all()
        # temp_records = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5,
        #     ConcTemp.temp6, ConcTemp.datetime).join(ConcNode, ConcNode.id == ConcTemp.conc_node_id).join(
        #     ConcGateway, ConcGateway.id == ConcTemp.conc_gateway_id).filter(and_(ConcGateway.gateway_addr == gatewayAddr,
        #     ConcNode.node_addr == nodeAddr, ConcTemp.datetime.between(startTime, endTime))).order_by(ConcTemp.datetime.desc()).all()

        temp_log = []

        for i in xrange(0,len(temp_records)):
            temp_log.append({"key": i,"image":"http://dummyimage.com/48x48/{1}/757575.png&text={0}".format(temp_records[i][7],index_color(1)[1:]),
                "conc_node_id": temp_records[i][8], "datetime": temp_records[i][7].strftime("%Y-%m-%d %H:%M:%S"), "temp1": temp_records[i][0], 
                "temp2": temp_records[i][1], "temp3": temp_records[i][2], "temp4": temp_records[i][3], "temp5": temp_records[i][4], 
                "temp6": temp_records[i][5], "battery_vol":temp_records[i][6]})


        temp_records_hour = []
        for j in xrange(0,len(datetime_list)):
            dt = datetime_list[j]
            print('---------dt.hour--------')
            print(dt.hour)
            temp_record_hour = []
            for i in xrange(0,len(temp_records)):
                if temp_records[i][7].hour == dt.hour:
                    print('-----temp_records[i]-----')
                    print(temp_records[i])
                    temp_record_hour.append(temp_records[i])

            temp_records_hour.append(temp_record_hour)



        print('--------temp_records_hour----------')
        print(temp_records_hour)

        wb = load_workbook('concrete_template.xlsx')
        ws = wb.active

        for i in xrange(0,len(temp_records_hour)):
            temp_to_insert = temp_records_hour[i]
            if temp_to_insert:
                temp1 = temp_to_insert[0][0]
                temp2 = temp_to_insert[0][1]
                temp3 = temp_to_insert[0][2]
                month = temp_to_insert[0][7].month
                day = temp_to_insert[0][7].day
                hour = temp_to_insert[0][7].hour
                nodeAddr = temp_to_insert[0][8]

                ws.cell(row=3*i+8, column=9, value=temp1)
                ws.cell(row=3*i+9, column=9, value=temp2)
                ws.cell(row=3*i+10, column=9, value=temp3)
                ws.cell(row=3*i+8, column=2, value=month)
                ws.cell(row=3*i+8, column=3, value=day)
                ws.cell(row=3*i+8, column=4, value=hour)

        wb.save(str(nodeAddr) + '_' + startTime0 + '_' + endTime0 + '_' "sample.xlsx")

        temps_reverse = temp_log[::-1]

        print('------------temps_records--------------')
        # print(temps_reverse)

        temps_record_dict = {"concTempRecord": temp_log}
        return temps_record_dict

    def delete(self, todo_id):
        pass

    def put(self, todo_id):
        pass


class ConcHistoryRecord(Resource):
    '''
        get the temp records by the input datetime. %H:%M:S%
    '''
    def get(self):

        parser = reqparse.RequestParser()
        parser.add_argument('gatewayAddr', type=str)
        parser.add_argument('nodeAddr', type=str)
        parser.add_argument('startTime', type=str)
        parser.add_argument('endTime', type=str)

        args = parser.parse_args()

        print('-------ConcTempRecord args---------', args)

        gatewayAddr = args['gatewayAddr']
        nodeAddr = args['nodeAddr']
        startTime0 = args['startTime']
        endTime0 = args['endTime']


        startTime = datetime.datetime.strptime(startTime0, "%Y-%m-%d %H:%M:%S")
        endTime = datetime.datetime.strptime(endTime0, "%Y-%m-%d %H:%M:%S")
        print('-------startTime--------')
        print('--------endTime-------')
        print(startTime)
        print(endTime)
        temp_records = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5,
            ConcTemp.temp6, ConcTemp.battery_vol, ConcTemp.datetime,ConcTemp.conc_node_id).filter(and_(ConcTemp.conc_node_id == nodeAddr, ConcTemp.datetime.between(startTime, endTime))).order_by(ConcTemp.datetime.desc()).all()
        # temp_records = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5,
        #     ConcTemp.temp6, ConcTemp.datetime).join(ConcNode, ConcNode.id == ConcTemp.conc_node_id).join(
        #     ConcGateway, ConcGateway.id == ConcTemp.conc_gateway_id).filter(and_(ConcGateway.gateway_addr == gatewayAddr,
        #     ConcNode.node_addr == nodeAddr, ConcTemp.datetime.between(startTime, endTime))).order_by(ConcTemp.datetime.desc()).all()

        temp_log = []
        temp_records_hour = []
        for i in xrange(0,len(temp_records)):
            if temp_records[i][7].hour == temp_records[i-1][7].hour:
                temp_records_hour.append(temp_records[i])

            temp_log.append({"key": i,"image":"http://dummyimage.com/48x48/{1}/757575.png&text={0}".format(temp_records[i][7],index_color(1)[1:]),
                "conc_node_id": temp_records[i][8], "datetime": temp_records[i][7].strftime("%Y-%m-%d %H:%M:%S"), "temp1": temp_records[i][0], 
                "temp2": temp_records[i][1], "temp3": temp_records[i][2], "temp4": temp_records[i][3], "temp5": temp_records[i][4], 
                "temp6": temp_records[i][5], "battery_vol":temp_records[i][6]})

        temps_reverse = temp_log[::-1]

        print('------------temps_records--------------')
        print(temps_reverse)

        temps_record_dict = {"list": temp_log}
        return temps_record_dict

    def delete(self, todo_id):
        pass

    def put(self, todo_id):
        pass


class ConcTempHourRecord(Resource):
    '''
        get the temp records by the input datetime. %H:%M:S%
    '''
    def get(self):

        parser = reqparse.RequestParser()
        parser.add_argument('gatewayAddr', type=str)
        parser.add_argument('nodeAddr', type=str)
        parser.add_argument('startTime', type=str)
        parser.add_argument('endTime', type=str)

        args = parser.parse_args()

        print('-------ConcTempHourRecord args---------', args)

        gatewayAddr = args['gatewayAddr']
        nodeAddr = args['nodeAddr']
        startTime0 = args['startTime']
        endTime0 = args['endTime']




        startTime = datetime.datetime.strptime(startTime0, "%Y-%m-%d %H:%M:%S")
        endTime = datetime.datetime.strptime(endTime0, "%Y-%m-%d %H:%M:%S")
        print('-------startTime--------')
        print('--------endTime-------')
        print(startTime)
        print(endTime)

        select_hours = []

        temp_records = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5,
            ConcTemp.temp6, ConcTemp.battery_vol, ConcTemp.datetime,ConcTemp.conc_node_id).filter(and_(ConcTemp.conc_node_id == nodeAddr, ConcTemp.datetime.between(startTime, endTime))).order_by(ConcTemp.datetime.desc()).all()
        # temp_records = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.temp4, ConcTemp.temp5,
        #     ConcTemp.temp6, ConcTemp.datetime).join(ConcNode, ConcNode.id == ConcTemp.conc_node_id).join(
        #     ConcGateway, ConcGateway.id == ConcTemp.conc_gateway_id).filter(and_(ConcGateway.gateway_addr == gatewayAddr,
        #     ConcNode.node_addr == nodeAddr, ConcTemp.datetime.between(startTime, endTime))).order_by(ConcTemp.datetime.desc()).all()

        temp_log = []
        temp_records_hour = []
        for i in xrange(0,len(temp_records)):
            if temp_records[i][7].hour == temp_records[i-1][7].hour:
                temp_records_hour.append(temp_records[i])

            temp_log.append({"key": i,"image":"http://dummyimage.com/48x48/{1}/757575.png&text={0}".format(temp_records[i][7],index_color(1)[1:]),
                "conc_node_id": temp_records[i][8], "datetime": temp_records[i][7].strftime("%Y-%m-%d %H:%M:%S"), "temp1": temp_records[i][0], 
                "temp2": temp_records[i][1], "temp3": temp_records[i][2], "temp4": temp_records[i][3], "temp5": temp_records[i][4], 
                "temp6": temp_records[i][5], "battery_vol":temp_records[i][6]})

        temps_reverse = temp_log[::-1]

        print('------------temps_records--------------')
        print(temps_reverse)

        temps_record_dict = {"concTempRecord": temp_log}
        return temps_record_dict

    def delete(self, todo_id):
        pass

    def put(self, todo_id):
        pass



class ConcDashboard(Resource):
    def return_status(self,a,b,c):
        max_abc = max(a,b,c)
        print('max_abc:', max_abc)
        if max_abc < 60:
            return 1
        elif (60 <= max_abc) and (max_abc <= 70):
            return 2
        else:
            return 3

    def get(self):
        nodes = db.session.query(ConcNode.node_addr).order_by(ConcNode.node_addr.desc()).all()
        print("nodes are:", nodes)
        statuses = []
        for node in nodes:
            # todo: repalce geteway_addr
            temps = db.session.query(ConcTemp.temp1, ConcTemp.temp2, ConcTemp.temp3, ConcTemp.datetime).filter(
                ConcNode.node_addr == node[0]).order_by(
                ConcTemp.datetime.desc()).first()
            if temps:
                status = {"name":node[0]+u"号测温点","status":self.return_status(temps[0], temps[1], temps[2]),
                    "content":"上：{0}℃, 中：{1}℃, 下：{2}℃".format(str(temps[0]),str(temps[1]),str(temps[2])),
                    "avatar":"http://dummyimage.com/48x48/f279aa/757575.png&text={0}".format(node[0]),"date":datetime.datetime.strftime(temps[3], "%Y-%m-%d %H:%M:%S")}
                statuses.append(status)
            else:
                statuses = []
        conc_dash_dic = {"concDash":statuses}
        print("conc_dash_dic", conc_dash_dic)
        return conc_dash_dic


class Menus(Resource):
    
    def get(self):
        menus = [
          {
            'id': '1',
            'icon': 'laptop',
            'name': '主界面',
            'route': '/concrete',
          },
          {
            'id': '12',
            'bpid': '1',
            'name': '温度详情',
            'icon': 'bulb',
            'route': '/concdetail',
          },
          {
            'id': '2',
            'bpid': '1',
            'name': '历史记录',
            'icon': 'calendar',
            'route': '/concrete_history',
          },{
            'id': '5',
            'bpid': '1',
            'name': '图表报告',
            'icon': 'area-chart',
          },
          {
            'id': '51',
            'bpid': '5',
            'mpid': '5',
            'name': '线状图',
            'icon': 'line-chart',
            'route': '/chart/lineChart',
          },
          {
            'id': '52',
            'bpid': '5',
            'mpid': '5',
            'name': '柱状图',
            'icon': 'bar-chart',
            'route': '/chart/barChart',
          },
          {
            'id': '53',
            'bpid': '5',
            'mpid': '5',
            'name': '面积图',
            'icon': 'area-chart',
            'route': '/chart/areaChart',
          },
          {
            'id': '9',
            'bpid': '1',
            'name': '用户中心',
            'icon': 'team',
          },
          {
            'id': '91',
            'bpid': '9',
            'mpid': '9',
            'name': 'user',
            'icon': 'team',
            'route': '/user',
          },
          {
            'id': '92',
            'bpid': '9',
            'mpid': '9',
            'name': 'user detail',
            'icon': 'team',
            'route': '/user_center',
          },
          {
            'id': '7',
            'bpid': '1',
            'name': '系统设置',
            'icon': 'setting',
            'route': '/system_setting',
          },

          {
            'id': '21',
            'mpid': '-1',
            'bpid': '2',
            'name': 'User Detail',
            'route': '/user/:id',
          },
          {
            'id': '3',
            'bpid': '1',
            'name': 'Request',
            'icon': 'api',
            'route': '/request',
          },
          {
            'id': '4',
            'bpid': '1',
            'name': 'UI Element',
            'icon': 'camera-o',
          },
          {
            'id': '41',
            'bpid': '4',
            'mpid': '4',
            'name': 'IconFont',
            'icon': 'heart-o',
            'route': '/UIElement/iconfont',
          },
          {
            'id': '42',
            'bpid': '4',
            'mpid': '4',
            'name': 'DataTable',
            'icon': 'database',
            'route': '/UIElement/dataTable',
          },
          {
            'id': '43',
            'bpid': '4',
            'mpid': '4',
            'name': 'DropOption',
            'icon': 'bars',
            'route': '/UIElement/dropOption',
          },
          {
            'id': '44',
            'bpid': '4',
            'mpid': '4',
            'name': 'Search',
            'icon': 'search',
            'route': '/UIElement/search',
          },
          {
            'id': '45',
            'bpid': '4',
            'mpid': '4',
            'name': 'Editor',
            'icon': 'edit',
            'route': '/UIElement/editor',
          },
          {
            'id': '46',
            'bpid': '4',
            'mpid': '4',
            'name': 'layer (Function)',
            'icon': 'credit-card',
            'route': '/UIElement/layer',
          },
          
          {
            'id': '6',
            'bpid': '1',
            'name': 'Test Navigation',
            'icon': 'setting',
          },
          {
            'id': '61',
            'bpid': '6',
            'mpid': '6',
            'name': 'Test Navigation1',
            'route': '/navigation/navigation1',
          },
          {
            'id': '62',
            'bpid': '6',
            'mpid': '6',
            'name': 'Test Navigation2',
            'route': '/navigation/navigation2',
          },
          {
            'id': '621',
            'bpid': '62',
            'mpid': '62',
            'name': 'Test Navigation21',
            'route': '/navigation/navigation2/navigation1',
          },
          {
            'id': '622',
            'bpid': '62',
            'mpid': '62',
            'name': 'Test Navigation22',
            'route': '/navigation/navigation2/navigation2',
          },
        ]
        return menus

    def delete(self, todo_id):
        pass

    def put(self, todo_id):
        pass


# class Dashboard(Resource):
#     def get(self):
#         sales = []
#         for i in range(7):
#             sales.append({"name":i, "乒乓球":round(random.uniform(1000, 2000), 2), "台球": round(random.uniform(2000, 3000), 2), "保龄球": round(random.uniform(1000,2000), 2), "美格菲": round(random.uniform(2000,3000), 2)})
#         cpu = {"usage":150,"space":825,"cpu":88,"data":[{"cpu":21},{"cpu":49},{"cpu":41},{"cpu":52},{"cpu":48},{"cpu":50},{"cpu":24},{"cpu":50},{"cpu":79},{"cpu":35},{"cpu":21},{"cpu":47},{"cpu":59},{"cpu":26},{"cpu":44},{"cpu":43},{"cpu":20},{"cpu":57},{"cpu":68},{"cpu":73}]}
#         browser = [{"name":"Google Chrome","percent":43.3,"status":1},{"name":"Mozilla Firefox","percent":33.4,"status":2},{"name":"Apple Safari","percent":34.6,"status":3},{"name":"Internet Explorer","percent":12.3,"status":4},{"name":"Opera Mini","percent":3.3,"status":1},{"name":"Chromium","percent":2.53,"status":1}]
#         user = {"name":"zuiidea","email":"zuiiidea@.gmail.com","sales":3241,"sold":3556,"avatar":"http://tva4.sinaimg.cn/crop.0.0.996.996.180/6ee6a3a3jw8f0ks5pk7btj20ro0rodi0.jpg"}
#         completed = [{"name":2008,"Task complete":972,"Cards Complete":750},{"name":2009,"Task complete":968,"Cards Complete":510},{"name":2010,"Task complete":471,"Cards Complete":313},{"name":2011,"Task complete":478,"Cards Complete":649},{"name":2012,"Task complete":868,"Cards Complete":387},{"name":2013,"Task complete":979,"Cards Complete":336},{"name":2014,"Task complete":217,"Cards Complete":459},{"name":2015,"Task complete":752,"Cards Complete":723},{"name":2016,"Task complete":488,"Cards Complete":567},{"name":2017,"Task complete":963,"Cards Complete":879},{"name":2018,"Task complete":897,"Cards Complete":931},{"name":2019,"Task complete":948,"Cards Complete":969}]
#         comments = [{"name":"Lee","status":1,"content":"Qblrja zyftht tporjpytp rfktb qukvrbtg pndw gkiqt ngkydnb oqdjqxg wzyijql chcmfprd snomcgxxtg vzwy dbjdmpb qnvoynr qiigukgg.","avatar":"http://dummyimage.com/48x48/a1f279/757575.png&text=L","date":"2016-05-23 02:26:28"},{"name":"Lopez","status":2,"content":"Dqr emkgbgmd jujth wvahktvbem lld eqkhuwy zhvexnv xede els fqgnered ubtue wnrepbb rtevx ghjgwdis jwbnpwyqa xjgd.","avatar":"http://dummyimage.com/48x48/f279c5/757575.png&text=L","date":"2016-01-05 04:24:28"},{"name":"Young","status":3,"content":"Fvviqxk evu fqnehc ftmcz rgb spehkpd qfvbikd bnlk folewxp ndsvl tdxbs cgsgcwrm ycazmlahpb ciyvywnp jpoiqbhwwl rhvhkruth kyym.","avatar":"http://dummyimage.com/48x48/79e8f2/757575.png&text=Y","date":"2016-08-10 16:20:04"},{"name":"White","status":2,"content":"Wpok aqijwq cyybg lxoxj fmxgkqtr mgmw vsvjt rcbyjzk bigblg hww sjvbhytz dwndu wshy grlyx.","avatar":"http://dummyimage.com/48x48/f2d879/757575.png&text=W","date":"2016-09-02 05:47:21"},{"name":"Hernandez","status":3,"content":"Yrwjdrhpkv scl janjnvy prwevmw nhmvi qusrg fkvkqg emmbdtg pkhmql tabuvosvz jfvsykwmz asfob zuou nqhre dhy rlbdufycjw.","avatar":"http://dummyimage.com/48x48/b579f2/757575.png&text=H","date":"2016-07-01 17:11:16"}]
#         recentSales = [{"id":1,"name":"Hernandez","status":3,"price":15.45,"date":"2016-08-05 01:48:34"},{"id":2,"name":"Anderson","status":4,"price":197.28,"date":"2016-04-12 23:18:26"},{"id":3,"name":"Harris","status":3,"price":90.24,"date":"2015-05-13 23:57:47"},{"id":4,"name":"Thomas","status":1,"price":162.9,"date":"2016-02-11 22:16:03"},{"id":5,"name":"Jones","status":4,"price":196.1,"date":"2015-05-03 18:16:08"},{"id":6,"name":"Anderson","status":4,"price":119.96,"date":"2016-04-28 12:46:07"},{"id":7,"name":"Clark","status":3,"price":109.33,"date":"2015-06-21 12:12:07"},{"id":8,"name":"Robinson","status":2,"price":48.4,"date":"2016-03-20 09:39:52"},{"id":9,"name":"Harris","status":1,"price":137.5,"date":"2016-03-07 09:29:20"},{"id":10,"name":"Martin","status":2,"price":43.2,"date":"2015-08-16 11:51:33"},{"id":11,"name":"Gonzalez","status":2,"price":72.2,"date":"2015-10-24 07:21:11"},{"id":12,"name":"Miller","status":2,"price":193.82,"date":"2015-03-12 21:09:06"},{"id":13,"name":"Thompson","status":3,"price":180.34,"date":"2015-01-18 22:07:41"},{"id":14,"name":"Lee","status":4,"price":46.03,"date":"2016-12-28 01:33:17"},{"id":15,"name":"Harris","status":1,"price":73.76,"date":"2015-06-30 11:10:08"},{"id":16,"name":"Smith","status":2,"price":116.6,"date":"2016-07-15 05:40:46"},{"id":17,"name":"Martinez","status":2,"price":110.66,"date":"2016-02-23 17:21:31"},{"id":18,"name":"Smith","status":2,"price":52.37,"date":"2015-03-12 14:19:55"},{"id":19,"name":"Thompson","status":4,"price":198.3,"date":"2015-04-07 04:48:27"},{"id":20,"name":"Johnson","status":2,"price":15.4,"date":"2016-05-06 07:10:59"},{"id":21,"name":"Martinez","status":3,"price":196.82,"date":"2015-06-24 13:26:42"},{"id":22,"name":"Miller","status":1,"price":157.3,"date":"2015-06-24 07:12:13"},{"id":23,"name":"Harris","status":3,"price":46.04,"date":"2015-09-27 17:55:38"},{"id":24,"name":"Rodriguez","status":1,"price":93.2,"date":"2016-06-21 02:41:56"},{"id":25,"name":"Martin","status":2,"price":82.48,"date":"2015-08-28 08:13:48"},{"id":26,"name":"Martin","status":2,"price":169.44,"date":"2015-11-28 21:21:17"},{"id":27,"name":"Brown","status":3,"price":46.2,"date":"2016-07-24 01:30:16"},{"id":28,"name":"Garcia","status":1,"price":102.39,"date":"2015-09-12 22:40:57"},{"id":29,"name":"Lewis","status":2,"price":182.7,"date":"2015-04-25 15:35:28"},{"id":30,"name":"Garcia","status":1,"price":59.69,"date":"2016-01-01 06:17:33"},{"id":31,"name":"Perez","status":4,"price":165.25,"date":"2015-09-15 16:15:51"},{"id":32,"name":"Walker","status":1,"price":16.84,"date":"2016-01-27 10:42:59"},{"id":33,"name":"Young","status":3,"price":20.9,"date":"2015-11-13 07:07:20"},{"id":34,"name":"Perez","status":4,"price":113.65,"date":"2016-04-12 11:58:06"},{"id":35,"name":"Hernandez","status":3,"price":89.6,"date":"2016-08-24 11:20:56"},{"id":36,"name":"Jones","status":4,"price":89.6,"date":"2016-05-27 03:03:16"}]
#         quote = {"name":"Joho Doe","title":"上海zonggonghui","content":"公告栏","avatar":"http://img.hb.aicdn.com/bc442cf0cc6f7940dcc567e465048d1a8d634493198c4-sPx5BR_fw236"}
#         numbers = [{"icon":"pay-circle-o","color":"#64ea91","title":"今日收入","number":round(random.uniform(4000, 5000), 2)},{"icon":"team","color":"#8fc9fb","title":"今日人数","number":random.randint(200,300)},{"icon":"message","color":"#d897eb","title":"今日新增会员","number":random.randint(20,30)},{"icon":"shopping-cart","color":"#f69899","title":"今日包场次数","number":random.randint(1,5)}]

#         return {'sales': sales, 'cpu': cpu, 'browser': browser, 'user': user, 'completed': completed, 'comments': comments, "recentSales": recentSales, "quote": quote, "numbers": numbers}

#     def delete(self):
#         pass

#     def put(self):
#         pass

#     def post(self):
#         pass


api.add_resource(TodoList, '/todos')
api.add_resource(Todo, '/todos/<todo_id>')
api.add_resource(Username, '/username/<username>')
api.add_resource(Menus, '/api/v1/menus')
api.add_resource(ConcRealtimeTemp, '/api/v1/concrete_realtime_temperature')
api.add_resource(ConcTemps, '/api/v1/concrete_temperatures')
api.add_resource(ConcTempRecord, '/api/v1/concrete_temperature_record')
api.add_resource(ConcTempHourRecord, '/api/v1/concrete_temperature_hour_record')
api.add_resource(ConcHistoryRecord, '/api/v1/concrete_history_record')

api.add_resource(ConcDashboard, '/api/v1/concrete_dashboard')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8888, debug=True)